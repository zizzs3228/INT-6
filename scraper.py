import fitz
from docx import Document
from openpyxl import load_workbook
import pytesseract
from PIL import Image
import io
import re
import os
import logging
import sys


def extract_text_from_pdf(file_path):
    # Открываем PDF-документ
    document = fitz.open(file_path)
    text = []

    for page_num in range(len(document)):
        page = document.load_page(page_num)
        
        page_text = page.get_text()
        if page_text.strip():
            text.append(page_text)
        else:
            pix = page.get_pixmap()
            img_data = io.BytesIO(pix.tobytes(output="png"))  # Сохранение изображения в формате PNG
            img = Image.open(img_data)
            ocr_text = pytesseract.image_to_string(img, lang='rus')  # 'rus+eng' для русского и английского языков
            text.append(ocr_text)

    return ' '.join(text)

def extract_text_from_docx(file_path:str)->str:
    doc = Document(file_path)
    text = []
    for paragraph in doc.paragraphs:
        text.append(paragraph.text)
    return ' '.join(text)

def extract_text_from_xlsx(file_path:str)->str:
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    text = []
    for row in sheet.iter_rows(values_only=True):
        row_text = " ".join([str(cell) if cell is not None else "" for cell in row])
        text.append(row_text)
    return ' '.join(text)

def text_preparation(text:str)->str:
    text = text.replace('\n',' ')
    while '  ' in text:
        text = text.replace('  ',' ')
    return text

def sensitive_data_finder(text:str)->str:
    email_pattern = r'([a-zA-Z0-9._-]+@[a-zA-Z0-9-]+\.[a-zA-Z]+)'
    phone_pattern = r'((\+7\d{10})|([^0-9](7|8)\d{10}[^0-9])|((7|8)\(\d{3}\)\d{7})|( \d{3} \d{2} \d{2} )|( \d{3}-\d{2}-\d{2} )|((7|8) \(\d{3}\) \d{3}-\d{2}-\d{2})|((7|8)-\d{3}-\d{3}-\d{2}-\d{2})|([^0-9]\d{7}[^0-9])|((7|8) \d{3} \d{3} \d{2} \d{2})|(\(\d{3}\) \d{3}-\d{2}-\d{2})|((7|8) \(\d{3}\) \d{3} \d{2} \d{2}))'
    company_name_pattern = re.compile(r'((ООО|ИП|АО|ПАО|НКО|ОП|ТСЖ|НАО|ЗАО|НПАО)( ?)(\"|\«| )[а-яА-Я0-9-_]+(\"|\»| ))', re.IGNORECASE)
    ul_inn_pattern = r'[^0-9]\d{10}[^0-9]'
    fl_inn_pattern = r'[^0-9]\d{12}[^0-9]'
    kpp_pattern = r'[^0-9]\d{9}[^0-9]'
    bik_pattern = r'[^0-9]04\d{7}[^0-9]'
    snils_pattern = r'\d{3}-\d{3}-\d{3} \d{2}'
    Full_FIO_pattern = r'[А-Я][а-я]+ [А-Я][а-я]+ [А-Я][а-я]+'
    Abr_FIO_patterns = r'([А-Я](\.|\. | )[А-Я](\.|\. | )[А-Я][а-я]+)'
    
    emails = re.findall(email_pattern,text)
    phones = re.findall(phone_pattern,text)
    company_names = re.findall(company_name_pattern,text)
    ul_inn = re.findall(ul_inn_pattern,text)
    fl_inn = re.findall(fl_inn_pattern,text)
    kpp = re.findall(kpp_pattern,text)
    bik = re.findall(bik_pattern,text)
    snilses = re.findall(snils_pattern,text)
    Full_FIOS = re.findall(Full_FIO_pattern,text)
    Abr_FIOS = re.findall(Abr_FIO_patterns,text)
    
    phones = [phone[0].strip() for phone in phones if phone[0] != '']
    company_names = [company_name[0].strip() for company_name in company_names if company_name[0] != '']
    ul_inn = [inn[1:-1].strip() for inn in ul_inn if inn != '']
    fl_inn = [inn[1:-1].strip() for inn in fl_inn if inn != '']
    kpp = [inn[1:-1].strip() for inn in kpp if inn != '' and not inn[1:-1].startswith('04')]
    bik = [inn[1:-1].strip() for inn in bik if inn != '']
    snilses = [snils.strip() for snils in snilses if snils != '']
    Full_FIOS = [FIO.strip() for FIO in Full_FIOS if FIO != '']
    Abr_FIOS = [FIO[0].strip() for FIO in Abr_FIOS if FIO != '']
    
    phones = list(set(phones))
    emails = list(set(emails))
    company_names = list(set(company_names))
    ul_inn = list(set(ul_inn))
    fl_inn = list(set(fl_inn))
    kpp = list(set(kpp))
    bik = list(set(bik))
    snilses = list(set(snilses))
    Full_FIOS = list(set(Full_FIOS))
    Abr_FIOS = list(set(Abr_FIOS))
    return phones, emails, company_names, ul_inn, fl_inn, kpp, bik, snilses, Full_FIOS, Abr_FIOS
    
    
if __name__=='__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[logging.FileHandler('output.log'), logging.StreamHandler()])
    
    
    path = "./files"
    list_dir = os.listdir(path)
    logging.info(f'Обработка файлов в директории: {path}')
    files = [file for file in list_dir if file.endswith('.docx') or file.endswith('.pdf') or file.endswith('.xlsx')]
    if not files:
        logging.info('В директории нет файлов для обработки')
        sys.exit(0)
    logging.info(f'Найдено файлов: {files}')
    logging.info('Подождите около минуты, пока обрабатываются файлы... Скорее всего, всё работает верно, просто OCR занятие не быстрое')
    logging.info('Если docker ps пустой, то выполнение завершено. Проверьте файл output.log для получения результатов.')
    
    for file in files:
        try:
            if file.endswith('.docx'):
                text = extract_text_from_docx(path+'/'+file)
            elif file.endswith('.pdf'):
                text = extract_text_from_pdf(path+'/'+file)
            elif file.endswith('.xlsx'):
                text = extract_text_from_xlsx(path+'/'+file)
        except Exception as e:
            logging.info(f'Ошибка при обработке файла: {file} - {e}')
            continue
        text = text_preparation(text)
        if text != '':
            phones, emails, company_names, ul_inn, fl_inn, kpp, bik, snilses, Full_FIOS, Abr_FIOS = sensitive_data_finder(text)
            logging.info(f'Файл: {file}')
            logging.info(f'Телефоны: {phones}')
            logging.info(f'Электронные почты: {emails}')
            logging.info(f'Названия компаний: {company_names}')
            logging.info(f'ИНН Юридических лиц: {ul_inn}')
            logging.info(f'ИНН Физических лиц: {fl_inn}')
            logging.info(f'КПП: {kpp}')
            logging.info(f'БИКи: {bik}')
            logging.info(f'СНИЛСы: {snilses}')
            logging.info(f'Полные ФИО: {Full_FIOS}')
            logging.info(f'Аббревиатуры ФИО: {Abr_FIOS}')
            logging.info('\n')
        else:
            logging.info(f'Файл: {file} - не содержит текста')
            logging.info('\n')
    